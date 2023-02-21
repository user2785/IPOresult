import pickle
import re
import urllib.request
import selenium.common.exceptions
import openpyxl
import os
import io
import base64
import CaptchaCracker as cc
import threading
import logging
logging.getLogger('tensorflow').disabled = True
import warnings

from silence_tensorflow import silence_tensorflow
from PIL import Image
from colorama import Fore, Style, init
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait

# Function to clear lines
def clear_line(n=1):
    LINE_UP = '\033[1A'
    LINE_CLEAR = '\x1b[2K'
    for i in range(n):
        print(LINE_UP, end=LINE_CLEAR)


# Function to Check Connection
def connection_check(host):
    try:
        urllib.request.urlopen(host)
    except:
        print("\nError: Connection Failed...")
        if browser != None and wb != None:
            close_all(0)  # Call close_all function
        exit()


# Function to open browser
def browser_open(url):
    global browser
    try:
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')
        browser = webdriver.Chrome(options=chrome_options)
        print("Chrome driver created successfully")
    except:
        try:
            edge_options = webdriver.EdgeOptions()
            edge_options.add_argument('-headless')
            browser = webdriver.Edge(options=edge_options)
            print("Edge driver created successfully")
        except:
            try:
                firefox_options = webdriver.FirefoxOptions()
                firefox_options.headless = True
                browser = webdriver.Firefox(options=firefox_options)
                print("Firefox driver created successfully")
            except:
                print("Error getting webdriver")
                print("run webdriver.py")
                close_all(1)  # Call close_all function

    browser.get(url)
    try:
        WebDriverWait(browser, 10).until(
            ec.presence_of_element_located(
                (
                    By.CSS_SELECTOR, '#companyShare > option:nth-child(4)')
            )
        )
    finally:
        test = 0


# Function to get Result
def check(client_boid, client_name, company_name):
    text = browser.find_element(By.ID, "boid")
    text.clear()
    text.send_keys(client_boid)
    result = browser.find_element(By.XPATH, "/html/body/app-root/app-allotment-result/div/div/div/div/form/div[3]").text
    if result == 'Invalid BOID':
        print(client_name + " : " + client_boid + " is incorrect. Correct before continue")
        close_all(0)  # Call close_all function
    result = get_captcha()  # Call get_captcha function
    if result == 'pass':
        return 0

    # Extracting allotted quantity
    find = re.compile(r"\d.*")
    allot = find.search(str(result))
    if allot is None:
        allotted = 0
    else:
        allotted = allot.group()
    units = print_result(allotted, client_name)  # Call print_result function
    xl_write(units, client_name, client_boid, company_name)  # Call xl_write function
    wb.save('results.xlsx')


# Function to save captcha temorarily
def save_img():
    image = browser.find_element(By.XPATH, '//img[@alt="captcha"]').get_attribute("src")
    recompile = re.compile(r'data:image/png;base64,.*', re.DOTALL)
    base64data = recompile.search(image)
    data1 = base64data.group()[22:]
    img = Image.open(io.BytesIO(base64.decodebytes(bytes(data1, "utf-8"))))
    img.save("captcha" + ".png")


# Function to Enter Captcha and to Check for error
def get_captcha():
    while True:
        captcha = browser.find_element(By.ID, "userCaptcha")
        save_img()  # Function to save captcha to file
        captcha_solved = auto_captcha_solver('captcha.png')  # auto decode captcha from tensorflow model
        os.remove('captcha.png')
        if len(captcha_solved)!= 5:
            continue
        captcha.click()
        captcha.send_keys(captcha_solved)         
        submit = browser.find_element(By.XPATH,
                    "/html/body/app-root/app-allotment-result/div/div/div/div/form/button")
        submit.click()
        result = browser.find_element(By.XPATH,
                    "/html/body/app-root/app-allotment-result/div/div/div/div/form/p[1]").text
        if result == '':
            result = browser.find_element(By.XPATH,
                    "/html/body/app-root/app-allotment-result/div/div/div/div/form/p[2]").text
        if result == '':
            result = browser.find_element(By.XPATH,
                    "/html/body/app-root/app-allotment-result/div/div/div/div/form/p[3]").text           
        if result[0:10] == "Connection":
            print("Error: " + result)
            close_all(2)  # Call close_all function
        if result[0:15] == "Invalid Captcha":
            continue
        if result == '':
            print("Something went wrong while checking result")
            close_all(1)  # Call close_all function
        else:
            return result


# Function to get captcha from trained model
def auto_captcha_solver(img_path):
    img_width = 150
    img_height = 40
    max_length = 5
    characters = {'1', '2', '3', '4', '5', '6', '7', '8', '9'}
    model_path = "Captcha_solve/model.h5"
    AM = cc.ApplyModel(model_path, img_width, img_height, max_length, characters)
    pred = AM.predict(img_path)
    clear_line(1)
    return pred


# Function to Print Result
def print_result(allotted, names):
    if allotted != 0:
        print(Fore.GREEN + Style.BRIGHT + " Congratulations, " + str(allotted) + " units is allotted for " + names)
    else:
        print(Fore.RED + Style.DIM + " Sorry, not allotted for " + names)
    return allotted


# Function to write obtained result in Excel
def xl_write(units, client_name, client_boid, company_name):
    maxvalue = sheet.max_row
    sheet['A' + str(maxvalue + 1)] = client_name
    sheet['B' + str(maxvalue + 1)] = client_boid
    sheet['C' + str(maxvalue + 1)] = int(units)
    sheet['D' + str(maxvalue + 1)] = company_name


# Function to close all running instances
def close_all(n):
    browser.quit()
    wb.close()
    exit(n)


# Start of the Code
url = "https://iporesult.cdsc.com.np"
silence_tensorflow()
warnings.filterwarnings("ignore", category=FutureWarning, module="numpy")
logging.disable(logging.CRITICAL)
logging.getLogger("cc").setLevel(logging.ERROR)
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
init(convert=True, autoreset=True)
#t1 = threading.Thread(target=auto_captcha_solver, args=('test.png',)) # Call auto_captcha_solver function to remove debug messeges
t2 = threading.Thread(target=connection_check, args=(url,))  # Call connection_check function
t3 = threading.Thread(target=browser_open, args=(url,)) # Call browser_open function

#t1.start()
t3.start()
t2.start()
wb = openpyxl.load_workbook('results.xlsx')
sheet = wb['Result']
t2.join()
t3.join()
#t1.join()
#os.system('clear||cls')
p,count = 999999, 0
while True:
    # Select Company, Check Result
    try:
        print("\n\t" + Fore.BLUE + Style.BRIGHT + browser.title + "\n")
        company = browser.find_element(By.ID, "companyShare")
        k = None
        option = None
        option = browser.find_elements(By.CSS_SELECTOR, "#companyShare > option", )
        for index, text in enumerate(option):
            print(f'{index} > {option[index].text}')
        sel = Select(company)
        
        while (int(p)<0 or int(p)>=len(option)):
            p = input("\nSelect company [or '0' to check all]:  ")
            clear_line(1)
            if p == '':
                close_all(0)  # Call close_all function
            continue
        p = int(p)
        if p == 0:
            check_all = range(1,len(option))
        elif p == None:
            close_all(0)  # Call close_all function
        else: 
            check_all = range(p,p+1)

        # Load BOID
        file = open("boid.pkl", "rb")
        data = pickle.load(file)
        file.close()

        for i in check_all:
            sel.select_by_index(i)
            selected_company = option[i].text
            print("\n\t" + Fore.BLUE + Style.BRIGHT + selected_company + '\n')

            # Check Result
            for boid, name in data.items():
                for row in range(2, sheet.max_row + 1):
                    if sheet['B' + str(row)].value == boid and sheet['D' + str(row)].value == selected_company:
                        print_result(sheet['C' + str(row)].value, sheet['A' + str(row)].value)  # Call print_result function
                        break

                else:
                    check(boid, name, selected_company)  # Call check function

        z = input("\npress 0 to check again  ")
        if str(z) == '0':
            p = 9999999
            continue
        break

    except selenium.common.exceptions.NoSuchWindowException:
        print(Fore.RED + "\n Error: NoSuchWindowException Exiting.... ")
    except selenium.common.exceptions.InvalidSessionIdException:
        print(Fore.RED + "\n Error: InvalidSessionIDException Exiting....")
    except selenium.common.exceptions.WebDriverException:
        print(Fore.RED + "\n Error: WebDriverException Something went wrong...")
        if count >=5:
            print("something is wrong with webdriver... Try again")
        else:
            count = count + 1
            os.system('clear||cls')
            continue
    except KeyboardInterrupt:
        print(Fore.RED + "\n Error: KeyboardInterrupt")
    close_all(0)  # Call close_all function
