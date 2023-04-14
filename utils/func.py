from pickle import load as lo
import openpyxl
import urllib.request as urllib_request
from colorama import Fore, Style
import re

def load_boid(filename):
    with open(filename, "rb") as file:
        data = lo(file)
    return data

# Function to clear lines
def clear_line(n=1):
    LINE_UP = '\033[1A'
    LINE_CLEAR = '\x1b[2K'
    for i in range(n):
        print(LINE_UP, end=LINE_CLEAR)
    

def import_excel(filename): 
    global wb
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Result']
    return sheet
    

# Function to write obtained result in Excel
def xl_write(units, client_name, client_boid, company_name, sheet):
    maxvalue = sheet.max_row
    sheet['A' + str(maxvalue + 1)] = client_name
    sheet['B' + str(maxvalue + 1)] = client_boid
    sheet['C' + str(maxvalue + 1)] = units
    sheet['D' + str(maxvalue + 1)] = company_name
    wb.save('data/results.xlsx')

    
def check_in_excel(sheet, boid, selected_company):
    for row in range(2, sheet.max_row + 1):
        if sheet['B' + str(row)].value == boid and sheet['D' + str(row)].value == selected_company:
            print_result(sheet['C' + str(row)].value, sheet['A' + str(row)].value)  # Call print_result function
            return True
    else:
        return False
    
    
# Function to Print Result
def print_result(allotted, name):
    if allotted != 0:
        print(f"{Fore.GREEN}{Style.BRIGHT} Congratulations, {allotted} units is allotted for {name}")
    else:
        print(f"{Fore.RED}{Style.DIM} Sorry, not allotted for {name}")


def re_compile(result):
    # Extracting allotted quantity
    find = re.compile(r"\d.*")
    allot = find.search(str(result))
    if allot is None:
        allotted = 0
    else:
        allotted = allot.group()
    return allotted


def close_wb():
    wb.save('data/results.xlsx')
    wb.close()