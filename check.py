import pickle
import re
import urllib.request
import selenium.common.exceptions
import openpyxl

from colorama import Fore, Style, init
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait


# Function to Check Connection
def connection_check(host):
    try:
        urllib.request.urlopen(host)
    except:
        print("\nError: Connection Failed...")
        exit()


# Function to wait for site to load
def wait():
    try:
        WebDriverWait(browser, 10).until(
            ec.presence_of_element_located(
                (
                    By.CSS_SELECTOR, '#companyShare > option:nth-child(4)')
            )
        )
    finally:
        print("\n\t" + Fore.BLUE + Style.BRIGHT + browser.title + "\n")


# Function to write obtained result in Excel
def xl_write(units, client_name, client_boid, company_name):
    maxvalue = sheet.max_row
    sheet['A' + str(maxvalue + 1)] = client_name
    sheet['B' + str(maxvalue + 1)] = client_boid
    sheet['C' + str(maxvalue + 1)] = int(units)
    sheet['D' + str(maxvalue + 1)] = company_name


# Function to get Result
def check(client_boid, client_name, company_name):
    text = browser.find_element(By.ID, "boid")
    text.clear()
    text.send_keys(client_boid)
    result = browser.find_element(By.XPATH, "/html/body/app-root/app-allotment-result/div/div/div/div/form/div[3]").text
    if result == 'Invalid BOID':
        print(client_name + " : " + client_boid + " is incorrect. Correct before continue")
        browser.quit()
        exit()
    result = get_captcha(client_name)  # Call get_captcha function
    if result == 'pass':
        return 0

    # Extracting allotted quantity
    find = re.compile(r"\d.*")
    allot = find.search(str(result))
    if allot is None:
        allotted = 0
    else:
        allotted = allot.group()
    units = print_result(allotted, client_name)  # Call print_result function
    xl_write(units, client_name, client_boid, company_name)  # Call xl_write function
    wb.save('results.xlsx')


# Function to Print Result
def print_result(allotted, names):
    if allotted != 0:
        print(Fore.GREEN + Style.BRIGHT + " Congratulations, " + str(allotted) + " units is allotted for " + names)
    else:
        print(Fore.RED + Style.DIM + " Sorry, not allotted for " + names)
    return allotted


# Function to Enter Captcha and to Check for error
def get_captcha(name_passed):
    while True:
        captcha = browser.find_element(By.ID, "userCaptcha")
        captcha.click()
        while True:
            if len(captcha.get_attribute("value")) == 5:
                if captcha.get_attribute("value") == '00000':
                    print(Fore.CYAN + "Skipped for " + name_passed)
                    captcha.clear()
                    return 'pass'
                submit = browser.find_element(By.XPATH,
                                              "/html/body/app-root/app-allotment-result/div/div/div/div/form/button")
                submit.click()
                break
        result = browser.find_element(By.XPATH,
                                      "/html/body/app-root/app-allotment-result/div/div/div/div/form/p[1]").text
        if result == '':
            result = browser.find_element(By.XPATH,
                                          "/html/body/app-root/app-allotment-result/div/div/div/div/form/p[2]").text
        if result[0:10] == "Connection":
            print("Error: " + result)
            browser.quit()
            exit(2)
        if result[0:15] == "Invalid Captcha":
            continue
        if result == '':
            print("Something went wrong while checking result")
            exit(1)
        else:
            return result


# Start of the Code
# initialize
init(convert=True, autoreset=True)
url = "https://iporesult.cdsc.com.np"
print(" Opening " + url)
print(" .......Please Wait.......  ")
connection_check(url)  # Call connection_check function

browser = webdriver.Firefox()
browser.set_window_size(500, 700)
handle_of_the_window = browser.current_window_handle
browser.minimize_window()
browser.get(url)
clear_line = '\033[A                                                                                                ' \
             '     \033[A '
wb = openpyxl.load_workbook('results.xlsx')
sheet = wb['Result']

wait()  # Call wait function

while True:
    # Select Company, Check Result
    try:
        company = browser.find_element(By.ID, "companyShare")
        k = None
        option = None
        try:
            for i in range(100):
                option = browser.find_elements(
                    By.CSS_SELECTOR, "#companyShare > option",
                )
                print(str(i) + "> " + option[i].text)
                k = i
        except IndexError:
            print("\n")

        sel = Select(company)
        while True:
            try:
                p = int(input("Select company:  "))
                if p > k or p <= 0:
                    print(clear_line)
                else:
                    break
            except ValueError:
                print(clear_line)

        sel.select_by_index(p)
        selected_company = option[p].text
        print("\n\t" + Fore.BLUE + Style.BRIGHT + selected_company + "\n")
        
        # Load BOID
        file = open("boid.pkl", "rb")
        data = pickle.load(file)
        file.close()

        # Check Result
        for boid, name in data.items():
            for row in range(2, sheet.max_row + 1):
                if sheet['B' + str(row)].value == boid and sheet['D' + str(row)].value == selected_company:
                    print_result(sheet['C' + str(row)].value, sheet['A' + str(row)].value)  # Call print_result function
                    break

            else:
                browser.switch_to.window(handle_of_the_window)
                browser.set_window_rect(0, 0)
                check(boid, name, selected_company)  # Call check function
        print('\n')
        browser.minimize_window()
        z = input("press 0 to check again  ")
        if str(z) == '0':
            continue
        break

    except selenium.common.exceptions.NoSuchWindowException:
        print(Fore.RED + "\n Error: NoSuchWindowException Exiting.... ")
    except selenium.common.exceptions.InvalidSessionIdException:
        print(Fore.RED + "\n Error: InvalidSessionIDException Exiting....")
    except selenium.common.exceptions.WebDriverException:
        print(Fore.RED + "\n Error: WebDriverException Something went wrong...")
    except KeyboardInterrupt:
        print(Fore.RED + "\n Error: KeyboardInterrupt")
    break
wb.close()
browser.quit()
exit()
